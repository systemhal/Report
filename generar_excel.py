import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_import_calculator():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # Estilos de Diseño
    # -------------------------------------------------------------
    font_family = "Segoe UI"
    
    # Fuentes
    title_font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=11, bold=True, color="C00000") # Rojo Marca
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=10, bold=True, color="000000")
    regular_font = Font(name=font_family, size=10, color="000000")
    small_gray_font = Font(name=font_family, size=9, italic=True, color="595959")
    
    # Rellenos
    navy_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid") # Negro Pizarra
    steel_blue_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # Rojo Marca
    section_fill = PatternFill(start_color="F9EBEA", end_color="F9EBEA", fill_type="solid") # Rojo claro / rosado
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo suave para entradas
    zebra_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid") # Gris muy claro
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Gris neutral
    accent_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Verde suave para resultados clave
    
    # Bordes
    thin_border_side = Side(style='thin', color='D3D3D3')
    thick_bottom_side = Side(style='double', color='000000')
    thin_top_side = Side(style='thin', color='000000')
    
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_total = Border(top=thin_top_side, bottom=thick_bottom_side, left=thin_border_side, right=thin_border_side)
    
    # Alineación
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_title = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Formatos de número
    format_usd = "$#,##0.00"
    format_pen = "S/ #,##0.00"
    format_percent = "0.0%"
    format_weight = '#,##0.00" kg"'
    format_volume = '#,##0.00" CBM"'
    format_qty = "#,##0"

    # =============================================================
    # HOJA 1: Parámetros y Tarifas
    # =============================================================
    ws1 = wb.active
    ws1.title = "1. Parametros y Tarifas"
    ws1.views.sheetView[0].showGridLines = True
    
    # Título Principal
    ws1.merge_cells("A1:C1")
    title_cell = ws1["A1"]
    title_cell.value = "CONFIGURACIÓN DE PARÁMETROS - IMPORTACIONES PERÚ"
    title_cell.font = title_font
    title_cell.fill = navy_fill
    title_cell.alignment = align_title
    ws1.row_dimensions[1].height = 40
    
    # Sección 1: Impuestos
    ws1.merge_cells("A3:C3")
    ws1["A3"] = "1. IMPUESTOS ADUANEROS (SUNAT)"
    ws1["A3"].font = section_font
    ws1["A3"].fill = section_fill
    ws1["A3"].alignment = align_left
    ws1.row_dimensions[3].height = 24
    
    tax_params = [
        ("Tasa IGV (Impuesto General a las Ventas)", 0.16, "Crédito fiscal (tasa estándar 16%)"),
        ("Tasa IPM (Impuesto Promoción Municipal)", 0.02, "Crédito fiscal (tasa estándar 2%)"),
        ("IGV + IPM Total (Base para importaciones)", "=B4+B5", "Suma total de impuestos a las ventas (18%)"),
        ("Percepción IGV (Primera Importación)", 0.10, "Aplicable en la primera DUA/DAM del importador"),
        ("Percepción IGV (Importador Frecuente)", 0.035, "Aplicable para importaciones habituales (defecto)"),
        ("Percepción IGV (Bienes Usados)", 0.05, "Aplicable para la importación de mercancías usadas"),
        ("Tipo de Cambio Oficial (S/ por USD)", 3.78, "Tasa de cambio comercial promedio (editable)")
    ]
    
    row_idx = 4
    for label, val, note in tax_params:
        ws1.cell(row=row_idx, column=1, value=label).font = regular_font
        ws1.cell(row=row_idx, column=1).alignment = align_left
        
        val_cell = ws1.cell(row=row_idx, column=2, value=val)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        
        if label.startswith("IGV + IPM"):
            val_cell.number_format = format_percent
        elif "Percepción" in label:
            val_cell.number_format = format_percent
            val_cell.fill = input_fill
        elif "Tipo de Cambio" in label:
            val_cell.number_format = format_pen
            val_cell.fill = input_fill
        else:
            val_cell.number_format = format_percent
            val_cell.fill = input_fill
            
        ws1.cell(row=row_idx, column=3, value=note).font = small_gray_font
        ws1.cell(row=row_idx, column=3).alignment = align_left
        
        for c in range(1, 4):
            ws1.cell(row=row_idx, column=c).border = border_cell
        ws1.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    # Sección 2: Gastos Locales
    row_idx += 1
    ws1.merge_cells(f"A{row_idx}:C{row_idx}")
    ws1[f"A{row_idx}"] = "2. GASTOS LOCALES DE DESPACHO (USD)"
    ws1[f"A{row_idx}"].font = section_font
    ws1[f"A{row_idx}"].fill = section_fill
    ws1[f"A{row_idx}"].alignment = align_left
    ws1.row_dimensions[row_idx].height = 24
    row_idx += 1
    
    start_locales = row_idx
    local_params = [
        ("Comisión Agente de Aduanas (%)", 0.005, "Porcentaje sobre el valor CIF de la importación"),
        ("Tarifa Mínima Agente de Aduanas (USD)", 200.00, "Cobro mínimo si la comisión por % es menor"),
        ("Gastos Portuarios / Desconsolidación", 180.00, "Gastos de descarga y desconsolidación de contenedor"),
        ("Almacenaje Aduanero (Depósito Temporal)", 150.00, "Almacén temporal autorizado de aduanas"),
        ("Flete Local (Transporte Terrestre Interno)", 220.00, "Traslado de aduana/puerto al almacén del cliente"),
        ("Gastos Bancarios (SWIFT / Transferencias)", 45.00, "Gastos de transferencia internacional al proveedor"),
        ("Gastos Administrativos / Logísticos Varios", 80.00, "Gastos operativos, copias, mensajería, etc."),
    ]
    
    for label, val, note in local_params:
        ws1.cell(row=row_idx, column=1, value=label).font = regular_font
        ws1.cell(row=row_idx, column=1).alignment = align_left
        
        val_cell = ws1.cell(row=row_idx, column=2, value=val)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.fill = input_fill
        
        if "%" in label:
            val_cell.number_format = format_percent
        else:
            val_cell.number_format = format_usd
            
        ws1.cell(row=row_idx, column=3, value=note).font = small_gray_font
        ws1.cell(row=row_idx, column=3).alignment = align_left
        
        for c in range(1, 4):
            ws1.cell(row=row_idx, column=c).border = border_cell
        ws1.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    # Fila de Total Gastos Fijos
    ws1.cell(row=row_idx, column=1, value="Total Gastos Locales Fijos").font = bold_font
    ws1.cell(row=row_idx, column=1).alignment = align_left
    total_locales_cell = ws1.cell(row=row_idx, column=2, value=f"=SUM(B{start_locales+2}:B{row_idx-1})")
    total_locales_cell.font = bold_font
    total_locales_cell.alignment = align_right
    total_locales_cell.number_format = format_usd
    total_locales_cell.fill = total_fill
    ws1.cell(row=row_idx, column=3, value="Suma de gastos portuarios, flete interno, almacén y varios").font = small_gray_font
    for c in range(1, 4):
        ws1.cell(row=row_idx, column=c).border = Border(top=thin_top_side, bottom=thin_top_side, left=thin_border_side, right=thin_border_side)
    ws1.row_dimensions[row_idx].height = 20
    
    # Sección 3: Logística
    row_idx += 2
    ws1.merge_cells(f"A{row_idx}:C{row_idx}")
    ws1[f"A{row_idx}"] = "3. FLETES Y SEGUROS DE REFERENCIA (USD)"
    ws1[f"A{row_idx}"].font = section_font
    ws1[f"A{row_idx}"].fill = section_fill
    ws1[f"A{row_idx}"].alignment = align_left
    ws1.row_dimensions[row_idx].height = 24
    row_idx += 1
    
    log_params = [
        ("Flete Marítimo por CBM (LCL)", 110.00, "Costo promedio de flete por metro cúbico"),
        ("Flete Marítimo por Tonelada LCL", 130.00, "Costo promedio de flete por 1000 kg"),
        ("Flete Aéreo por Kilogramo", 4.80, "Tarifa de flete aéreo internacional por kg"),
        ("Seguro Estimado de Aduana (%)", 0.015, "1.5% del valor FOB + Flete (fórmula de tabla SUNAT)")
    ]
    
    for label, val, note in log_params:
        ws1.cell(row=row_idx, column=1, value=label).font = regular_font
        ws1.cell(row=row_idx, column=1).alignment = align_left
        
        val_cell = ws1.cell(row=row_idx, column=2, value=val)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.fill = input_fill
        
        if "%" in label:
            val_cell.number_format = format_percent
        else:
            val_cell.number_format = format_usd
            
        ws1.cell(row=row_idx, column=3, value=note).font = small_gray_font
        ws1.cell(row=row_idx, column=3).alignment = align_left
        
        for c in range(1, 4):
            ws1.cell(row=row_idx, column=c).border = border_cell
        ws1.row_dimensions[row_idx].height = 20
        row_idx += 1

    # Ajustar anchos
    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 50

    # =============================================================
    # HOJA 2: Calculadora Individual
    # =============================================================
    ws2 = wb.create_sheet(title="2. Calculadora Individual")
    ws2.views.sheetView[0].showGridLines = True
    
    # Título Principal
    ws2.merge_cells("A1:C1")
    title_cell2 = ws2["A1"]
    title_cell2.value = "CALCULADORA DE IMPORTACIÓN INDIVIDUAL"
    title_cell2.font = title_font
    title_cell2.fill = navy_fill
    title_cell2.alignment = align_title
    ws2.row_dimensions[1].height = 40
    
    # Datos de Entrada
    ws2.merge_cells("A3:C3")
    ws2["A3"] = "DATOS DE LA MERCANCÍA Y EMBARQUE"
    ws2["A3"].font = section_font
    ws2["A3"].fill = section_fill
    ws2["A3"].alignment = align_left
    ws2.row_dimensions[3].height = 24
    
    inputs = [
        ("Nombre del Producto", "Audífonos Bluetooth Premium", "Identificador comercial", "@"),
        ("Precio FOB Total (USD)", 3500.00, "Valor de la mercancía en puerto de origen", format_usd),
        ("Cantidad de Unidades (Uds)", 500, "Número de productos a importar", format_qty),
        ("Peso Neto Total (kg)", 85.0, "Peso neto total de la carga", format_weight),
        ("Volumen Total (CBM)", 0.65, "Volumen en metros cúbicos", format_volume),
        ("Tipo de Transporte", "Marítimo", "Seleccionar 'Marítimo' o 'Aéreo'", None),
        ("Origen de la Mercancía", "China", "Seleccionar origen para recargos de flete", None),
        ("Tasa Ad-Valorem / Arancel (%)", 0.06, "Tasa SUNAT según partida (0%, 4%, 6%, 11%)", format_percent),
        ("Percepción IGV Aplicable", "Importador Frecuente (3.5%)", "Tasa de percepción de IGV", None),
        ("Margen de Ganancia Deseado (%)", 0.35, "Margen de ganancia bruto objetivo", format_percent)
    ]
    
    row_idx = 4
    for label, val, note, num_fmt in inputs:
        ws2.cell(row=row_idx, column=1, value=label).font = regular_font
        ws2.cell(row=row_idx, column=1).alignment = align_left
        
        val_cell = ws2.cell(row=row_idx, column=2, value=val)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.fill = input_fill
        
        if num_fmt:
            val_cell.number_format = num_fmt
        else:
            val_cell.alignment = align_center
            
        ws2.cell(row=row_idx, column=3, value=note).font = small_gray_font
        ws2.cell(row=row_idx, column=3).alignment = align_left
        
        for c in range(1, 4):
            ws2.cell(row=row_idx, column=c).border = border_cell
        ws2.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    # Data Validation para desplegables
    # Tipo de Transporte
    dv_transporte = DataValidation(type="list", formula1='"Marítimo,Aéreo"', allow_blank=True)
    ws2.add_data_validation(dv_transporte)
    dv_transporte.add(ws2["B9"]) # Fila 9 es Tipo de Transporte
    
    # Origen
    dv_origen = DataValidation(type="list", formula1='"China,EEUU,Europa,Otros"', allow_blank=True)
    ws2.add_data_validation(dv_origen)
    dv_origen.add(ws2["B10"]) # Fila 10 es Origen
    
    # Arancel
    dv_arancel = DataValidation(type="list", formula1='"0.0,0.04,0.06,0.11"', allow_blank=True)
    ws2.add_data_validation(dv_arancel)
    dv_arancel.add(ws2["B11"]) # Fila 11 es Arancel
    
    # Percepción
    dv_percepcion = DataValidation(type="list", formula1='"Primera Importación (10%),Importador Frecuente (3.5%),Bienes Usados (5%),No Aplicable (0%)"', allow_blank=True)
    ws2.add_data_validation(dv_percepcion)
    dv_percepcion.add(ws2["B12"]) # Fila 12 es Percepcion

    # Sección Desglose de Gastos
    row_idx += 1
    ws2.merge_cells(f"A{row_idx}:C{row_idx}")
    ws2[f"A{row_idx}"] = "DESGLOSE DE GASTOS E IMPUESTOS DE IMPORTACIÓN"
    ws2[f"A{row_idx}"].font = section_font
    ws2[f"A{row_idx}"].fill = section_fill
    ws2[f"A{row_idx}"].alignment = align_left
    ws2.row_dimensions[row_idx].height = 24
    row_idx += 1
    
    # FOB = B5, Qty = B6, Peso = B7, Vol = B8, Trans = B9, Origen = B10, Arancel = B11, Percep = B12, Margen = B13
    desglose = [
        ("Flete Internacional", '=IF(B9="Marítimo", MAX(B8*\'1. Parametros y Tarifas\'!$B$23, (B7/1000)*\'1. Parametros y Tarifas\'!$B$24), B7*\'1. Parametros y Tarifas\'!$B$25) * IF(B10="Europa", 1.2, IF(B10="EEUU", 1.1, 1.0))', "Prorrateado por peso o volumen más recargo de origen"),
        ("Seguro de Aduana (Estimado/Real)", '=(B5+B16)*\'1. Parametros y Tarifas\'!$B$26', "Costo de póliza o seguro estimado SUNAT"),
        ("Valor CIF (Costo + Flete + Seguro)", '=B5+B16+B17', "Valor base imponible de aduana (CIF)"),
        ("Derechos Arancelarios (Ad-Valorem)", '=B18*B11', "Arancel cobrado en aduana peruana"),
        ("Impuesto General a las Ventas (IGV 16%)", '=(B18+B19)*\'1. Parametros y Tarifas\'!$B$4', "Impuesto local IGV (Crédito Fiscal)"),
        ("Impuesto Promoción Municipal (IPM 2%)", '=(B18+B19)*\'1. Parametros y Tarifas\'!$B$5', "Impuesto local IPM (Crédito Fiscal)"),
        ("Gastos Operativos Locales (Despacho)", '=MAX(\'1. Parametros y Tarifas\'!$B$14, B18*\'1. Parametros y Tarifas\'!$B$13) + \'1. Parametros y Tarifas\'!$B$20', "Agente aduanas, puerto, almacén y transporte local"),
        ("Percepción IGV (SUNAT)", '=(B18+B19+B20+B21)*IF(B12="Primera Importación (10%)", \'1. Parametros y Tarifas\'!$B$7, IF(B12="Bienes Usados (5%)", \'1. Parametros y Tarifas\'!$B$9, IF(B12="No Aplicable (0%)", 0, \'1. Parametros y Tarifas\'!$B$8)))', "Pago por adelantado del IGV local (Crédito Fiscal)"),
    ]
    
    for label, formula, note in desglose:
        ws2.cell(row=row_idx, column=1, value=label).font = bold_font if "Valor CIF" in label else regular_font
        ws2.cell(row=row_idx, column=1).alignment = align_left
        
        val_cell = ws2.cell(row=row_idx, column=2, value=formula)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.number_format = format_usd
        
        if "Valor CIF" in label:
            val_cell.fill = total_fill
            
        ws2.cell(row=row_idx, column=3, value=note).font = small_gray_font
        ws2.cell(row=row_idx, column=3).alignment = align_left
        
        for c in range(1, 4):
            ws2.cell(row=row_idx, column=c).border = border_cell
        ws2.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    # Sección Análisis de Resultados
    row_idx += 1
    ws2.merge_cells(f"A{row_idx}:C{row_idx}")
    ws2[f"A{row_idx}"] = "ANÁLISIS DE COSTOS REALES Y PRECIOS DE VENTA"
    ws2[f"A{row_idx}"].font = section_font
    ws2[f"A{row_idx}"].fill = section_fill
    ws2[f"A{row_idx}"].alignment = align_left
    ws2.row_dimensions[row_idx].height = 24
    row_idx += 1
    
    resultados = [
        ("Costo Total en Almacén (Real Cost)", "=B18+B19+B22", "Suma de CIF + Arancel + Gastos Locales (No incluye IGV)"),
        ("Costo Unitario en Almacén (USD)", "=B26/B6", "Costo unitario real puesto en almacén (excl. IGV)"),
        ("Costo Unitario en Almacén (PEN)", "=B27*'1. Parametros y Tarifas'!$B$10", "Costo unitario real convertido a Soles"),
        ("Impuestos Recuperables (Crédito Fiscal)", "=B20+B21+B23", "IGV + IPM + Percepción que se recuperan en ventas locales"),
        ("Desembolso de Caja Total (Caja)", "=B26+B29", "Dinero total necesario para desaduanar y almacenar"),
        ("Precio Venta Sugerido (USD - Sin IGV)", "=B27/(1-B13)", "Precio unitario para obtener el margen neto deseado"),
        ("Precio Venta Sugerido (PEN - Sin IGV)", "=B31*'1. Parametros y Tarifas'!$B$10", "Precio unitario en Soles sin IGV local"),
        ("Precio Venta Sugerido Público (PEN - Con IGV)", "=B32*1.18", "Precio final al público con IGV incluido (18%)"),
        ("Utilidad Neta Total Estimada (USD)", "=(B31-B27)*B6", "Ganancia bruta total proyectada (excluyendo IGV)")
    ]
    
    # Note: FOB = B5, Qty = B6, Peso = B7, Vol = B8, Trans = B9, Origen = B10, Arancel = B11, Percep = B12, Margen = B14. 
    # Let's adjust cell references:
    # A4: Nombre (B4)
    # A5: FOB (B5)
    # A6: Qty (B6)
    # A7: Peso (B7)
    # A8: Vol (B8)
    # A9: Trans (B9)
    # A10: Origen (B10)
    # A11: Arancel (B11)
    # A12: Percep (B12)
    # A13: Margen (B13) -- Wait, the inputs list has 10 elements. Let's see:
    # index 4: Nombre (B4)
    # index 5: FOB (B5)
    # index 6: Qty (B6)
    # index 7: Peso (B7)
    # index 8: Vol (B8)
    # index 9: Trans (B9)
    # index 10: Origen (B10)
    # index 11: Arancel (B11)
    # index 12: Percep (B12)
    # index 13: Margen (B13)
    # So B6 is Qty, B7 is Peso, B8 is Vol, B9 is Trans, B10 is Origen, B11 is Arancel, B12 is Percep, B13 is Margen.
    # Therefore, in the formulas:
    # Flete = B16 -> =IF(B9="Marítimo", MAX(B8*'1. Parametros y Tarifas'!$B$23, (B7/1000)*'1. Parametros y Tarifas'!$B$24), B7*'1. Parametros y Tarifas'!$B$25) * IF(B10="Europa", 1.2, IF(B10="EEUU", 1.1, 1.0))
    # Seguro = B17 -> =(B5+B16)*'1. Parametros y Tarifas'!$B$26
    # CIF = B18 -> =B5+B16+B17
    # Arancel = B19 -> =B18*B11
    # IGV = B20 -> =(B18+B19)*'1. Parametros y Tarifas'!$B$4
    # IPM = B21 -> =(B18+B19)*'1. Parametros y Tarifas'!$B$5
    # Gastos Locales = B22 -> =MAX('1. Parametros y Tarifas'!$B$14, B18*'1. Parametros y Tarifas'!$B$13) + '1. Parametros y Tarifas'!$B$20
    # Percepción = B23 -> =(B18+B19+B20+B21)*IF(B12="Primera Importación (10%)", '1. Parametros y Tarifas'!$B$7, IF(B12="Bienes Usados (5%)", '1. Parametros y Tarifas'!$B$9, IF(B12="No Aplicable (0%)", 0, '1. Parametros y Tarifas'!$B$8)))
    
    # Now outcomes starts at Row 26:
    # Costo Almacen = B26 -> =B18+B19+B22
    # Cost Unit Alm USD = B27 -> =B26/B6
    # Cost Unit Alm PEN = B28 -> =B27*'1. Parametros y Tarifas'!$B$10
    # Impuestos Recup = B29 -> =B20+B21+B23
    # Desembolso Caja = B30 -> =B26+B29
    # P. Venta USD = B31 -> =B27/(1-B13)
    # P. Venta PEN = B32 -> =B31*'1. Parametros y Tarifas'!$B$10
    # P. Venta Pub PEN = B33 -> =B32*1.18
    # Utilidad = B34 -> =(B31-B27)*B6

    # That is perfectly correct! Let's write the code with these exact references.
    
    ws2_formulas = [
        ("Costo Total en Almacén (Real Cost)", "=B18+B19+B22", "Suma de CIF + Arancel + Gastos Locales (No incluye IGV)"),
        ("Costo Unitario en Almacén (USD)", "=B26/B6", "Costo unitario real puesto en almacén (excl. IGV)"),
        ("Costo Unitario en Almacén (PEN)", "=B27*'1. Parametros y Tarifas'!$B$10", "Costo unitario real convertido a Soles"),
        ("Impuestos Recuperables (Crédito Fiscal)", "=B20+B21+B23", "IGV + IPM + Percepción que se recuperan en ventas locales"),
        ("Desembolso de Caja Total (Caja)", "=B26+B29", "Dinero total necesario para desaduanar y almacenar"),
        ("Precio Venta Sugerido (USD - Sin IGV)", "=B27/(1-B13)", "Precio unitario para obtener el margen neto deseado"),
        ("Precio Venta Sugerido (PEN - Sin IGV)", "=B31*'1. Parametros y Tarifas'!$B$10", "Precio unitario en Soles sin IGV local"),
        ("Precio Venta Sugerido Público (PEN - Con IGV)", "=B32*1.18", "Precio final al público con IGV incluido (18%)"),
        ("Utilidad Neta Total Estimada (USD)", "=(B31-B27)*B6", "Ganancia bruta total proyectada (excluyendo IGV)")
    ]

    for label, formula, note in ws2_formulas:
        is_key = "Precio Venta" in label or "Costo Unitario" in label
        
        ws2.cell(row=row_idx, column=1, value=label).font = bold_font if is_key else regular_font
        ws2.cell(row=row_idx, column=1).alignment = align_left
        
        val_cell = ws2.cell(row=row_idx, column=2, value=formula)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        
        if "PEN" in label:
            val_cell.number_format = format_pen
        elif "Ad-Valorem" in label:
            val_cell.number_format = format_percent
        elif "Estimada" in label or "Total" in label or "USD" in label:
            val_cell.number_format = format_usd
        else:
            val_cell.number_format = format_usd
            
        if is_key:
            val_cell.fill = accent_green_fill
            
        ws2.cell(row=row_idx, column=3, value=note).font = small_gray_font
        ws2.cell(row=row_idx, column=3).alignment = align_left
        
        for c in range(1, 4):
            ws2.cell(row=row_idx, column=c).border = border_cell
        ws2.row_dimensions[row_idx].height = 20
        row_idx += 1

    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 50

    # =============================================================
    # HOJA 3: Prorrateo Multi-Producto
    # =============================================================
    ws3 = wb.create_sheet(title="3. Prorrateo Multi-Producto")
    ws3.views.sheetView[0].showGridLines = True
    
    # Título Principal
    ws3.merge_cells("A1:Y1")
    title_cell3 = ws3["A1"]
    title_cell3.value = "CUADRO DE PRORRATEO Y DISTRIBUCIÓN DE COSTOS DE IMPORTACIÓN"
    title_cell3.font = title_font
    title_cell3.fill = navy_fill
    title_cell3.alignment = align_title
    ws3.row_dimensions[1].height = 40
    
    # Entradas de la Factura General
    ws3.merge_cells("A3:D3")
    ws3["A3"] = "GASTOS GENERALES DEL EMBARQUE"
    ws3["A3"].font = section_font
    ws3["A3"].fill = section_fill
    ws3["A3"].alignment = align_left
    ws3.row_dimensions[3].height = 24
    
    # Entradas generales (B4 a B9)
    gen_inputs = [
        ("Flete Internacional Facturado (USD)", 1850.00, format_usd),
        ("Seguro Internacional Facturado (USD)", 125.00, format_usd),
        ("Gastos Locales de Despacho (USD)", 890.00, format_usd),
        ("Criterio de Prorrateo del Flete", "Volumen", None), # dropdown: Peso, Volumen, Valor FOB
        ("Tasa de Percepción IGV (%)", 0.035, format_percent),
        ("Tipo de Cambio Oficial (S/.)", "='1. Parametros y Tarifas'!B10", format_pen)
    ]
    
    r_idx = 4
    for lbl, val, fmt in gen_inputs:
        ws3.cell(row=r_idx, column=1, value=lbl).font = regular_font
        ws3.cell(row=r_idx, column=1).alignment = align_left
        
        val_cell = ws3.cell(row=r_idx, column=2, value=val)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.fill = input_fill
        
        if fmt:
            val_cell.number_format = fmt
        else:
            val_cell.alignment = align_center
            
        for c in range(1, 3):
            ws3.cell(row=r_idx, column=c).border = border_cell
        r_idx += 1
        
    # Agregar desplegable para el criterio de prorrateo
    dv_prorrateo = DataValidation(type="list", formula1='"Peso,Volumen,Valor FOB"', allow_blank=True)
    ws3.add_data_validation(dv_prorrateo)
    dv_prorrateo.add(ws3["B7"]) # Celda B7 es el Criterio

    # Headers de la Tabla de Productos en Fila 11
    headers = [
        "Item", "SKU / Código", "Descripción", "Cantidad (Uds)", "FOB Unit (USD)", 
        "Peso Unit (kg)", "Vol Unit (CBM)", "Arancel %", "FOB Total (USD)", "Peso Total (kg)", 
        "Vol Total (CBM)", "Flete Asig (USD)", "Seguro Asig (USD)", "CIF Total (USD)", "CIF Unit (USD)", 
        "Ad-Valorem (USD)", "Gastos Loc. (USD)", "Cost Almacén (USD)", "Cost Unit Alm (USD)", "Margen %", 
        "P. Venta USD (sin IGV)", "P. Venta PEN (sin IGV)", "Util Total (USD)", "IGV+IPM Aduana", "Percepción (USD)"
    ]
    
    ws3.row_dimensions[11].height = 28
    for col_idx, text in enumerate(headers, 1):
        cell = ws3.cell(row=11, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = steel_blue_fill
        cell.alignment = align_center
        cell.border = border_cell
        
    # Datos de Ejemplo (3 productos)
    sample_data = [
        (1, "SKU-9081", "Smartphones Android 12", 200, 150.00, 0.25, 0.002, 0.00, 0.35),
        (2, "SKU-4402", "Fundas de Silicona TPU", 1500, 1.20, 0.04, 0.0005, 0.06, 0.40),
        (3, "SKU-5211", "Cargadores Inalámbricos 15W", 800, 8.50, 0.12, 0.001, 0.06, 0.35)
    ]
    
    # Rellenar 15 filas (de la 12 a la 26) para que quepa un embarque real
    # Las primeras 3 filas tendrán datos, las otras vacías
    for r in range(12, 27):
        data_idx = r - 12
        is_even = (r % 2 == 0)
        row_fill = zebra_fill if is_even else PatternFill(fill_type=None)
        
        # Valores por defecto para fila vacía
        item_no = ""
        sku = ""
        desc = ""
        qty = ""
        fob_unit = ""
        peso_unit = ""
        vol_unit = ""
        arancel = ""
        margen = ""
        
        # Cargar datos si existen
        if data_idx < len(sample_data):
            item_no, sku, desc, qty, fob_unit, peso_unit, vol_unit, arancel, margen = sample_data[data_idx]
            
        ws3.cell(row=r, column=1, value=item_no).alignment = align_center
        ws3.cell(row=r, column=2, value=sku).alignment = align_center
        ws3.cell(row=r, column=3, value=desc).alignment = align_left
        
        c_qty = ws3.cell(row=r, column=4, value=qty)
        c_qty.alignment = align_right
        c_qty.number_format = format_qty
        c_qty.fill = input_fill if data_idx < len(sample_data) else row_fill
        
        c_fob = ws3.cell(row=r, column=5, value=fob_unit)
        c_fob.alignment = align_right
        c_fob.number_format = format_usd
        c_fob.fill = input_fill if data_idx < len(sample_data) else row_fill
        
        c_pes = ws3.cell(row=r, column=6, value=peso_unit)
        c_pes.alignment = align_right
        c_pes.number_format = format_weight
        c_pes.fill = input_fill if data_idx < len(sample_data) else row_fill
        
        c_vol = ws3.cell(row=r, column=7, value=vol_unit)
        c_vol.alignment = align_right
        c_vol.number_format = format_volume
        c_vol.fill = input_fill if data_idx < len(sample_data) else row_fill
        
        c_aran = ws3.cell(row=r, column=8, value=arancel)
        c_aran.alignment = align_right
        c_aran.number_format = format_percent
        c_aran.fill = input_fill if data_idx < len(sample_data) else row_fill
        
        # Fórmulas
        # FOB Total = Cantidad * FOB Unit
        ws3.cell(row=r, column=9, value=f"=D{r}*E{r}").number_format = format_usd
        # Peso Total = Cantidad * Peso Unit
        ws3.cell(row=r, column=10, value=f"=D{r}*F{r}").number_format = format_weight
        # Vol Total = Cantidad * Vol Unit
        ws3.cell(row=r, column=11, value=f"=D{r}*G{r}").number_format = format_volume
        
        # Flete Asignado
        # Depende de la opción en B7 (Peso, Volumen o Valor FOB)
        # Se calcula el prorrateo de los USD en B4
        flete_f = f'=IFERROR(IF($B$7="Peso", J{r}/SUM($J$12:$J$26)*$B$4, IF($B$7="Volumen", K{r}/SUM($K$12:$K$26)*$B$4, I{r}/SUM($I$12:$I$26)*$B$4)), 0)'
        ws3.cell(row=r, column=12, value=flete_f).number_format = format_usd
        
        # Seguro Asignado
        # Seguro se prorratea por el valor FOB total (I)
        seguro_f = f'=IFERROR(I{r}/SUM($I$12:$I$26)*$B$5, 0)'
        ws3.cell(row=r, column=13, value=seguro_f).number_format = format_usd
        
        # CIF Total = FOB + Flete + Seguro
        ws3.cell(row=r, column=14, value=f"=I{r}+L{r}+M{r}").number_format = format_usd
        # CIF Unit = CIF Total / Cantidad
        ws3.cell(row=r, column=15, value=f"=IFERROR(N{r}/D{r}, 0)").number_format = format_usd
        
        # Ad-Valorem = CIF Total * Arancel %
        ws3.cell(row=r, column=16, value=f"=N{r}*H{r}").number_format = format_usd
        
        # Gastos Locales
        # Se prorratea por valor FOB total (I) de los USD en B6
        gastos_f = f'=IFERROR(I{r}/SUM($I$12:$I$26)*$B$6, 0)'
        ws3.cell(row=r, column=17, value=gastos_f).number_format = format_usd
        
        # Costo Total en Almacén = CIF Total + AdValorem + Gastos Locales
        ws3.cell(row=r, column=18, value=f"=N{r}+P{r}+Q{r}").number_format = format_usd
        # Costo Unitario en Almacén = Costo Almacén / Cantidad
        ws3.cell(row=r, column=19, value=f"=IFERROR(R{r}/D{r}, 0)").number_format = format_usd
        ws3.cell(row=r, column=19).fill = accent_green_fill
        
        # Margen deseado %
        c_marg = ws3.cell(row=r, column=20, value=margen)
        c_marg.alignment = align_right
        c_marg.number_format = format_percent
        c_marg.fill = input_fill if data_idx < len(sample_data) else row_fill
        
        # Precio de Venta Unitario Sugerido (USD) = Cost Unit Alm / (1 - Margen)
        ws3.cell(row=r, column=21, value=f"=IFERROR(S{r}/(1-T{r}), 0)").number_format = format_usd
        ws3.cell(row=r, column=21).fill = accent_green_fill
        
        # Precio de Venta Unitario Sugerido (PEN) = Precio Venta USD * Tipo Cambio B9
        ws3.cell(row=r, column=22, value=f"=U{r}*$B$9").number_format = format_pen
        ws3.cell(row=r, column=22).fill = accent_green_fill
        
        # Utilidad Total = (Precio Venta USD - Cost Unit Alm USD) * Cantidad
        ws3.cell(row=r, column=23, value=f"=(U{r}-S{r})*D{r}").number_format = format_usd
        
        # IGV+IPM Aduana = (CIF Total + AdValorem) * (IGV + IPM de Config)
        igvipm_f = f"=(N{r}+P{r})*'1. Parametros y Tarifas'!$B$6"
        ws3.cell(row=r, column=24, value=igvipm_f).number_format = format_usd
        
        # Percepción = (CIF + AdValorem + IGV+IPM) * Percepcion % de B8
        percep_f = f"=(N{r}+P{r}+X{r})*$B$8"
        ws3.cell(row=r, column=25, value=percep_f).number_format = format_usd
        
        # Aplicar fuentes y bordes
        for c in range(1, 26):
            cell = ws3.cell(row=r, column=c)
            if c not in [4, 5, 6, 7, 8, 20]:  # No sobreescribir inputs
                cell.font = regular_font
                if c not in [19, 21, 22]: # Evitar sobreescribir rellenos de destaque
                    cell.fill = row_fill
            cell.border = border_cell
        ws3.row_dimensions[r].height = 20

    # Fila de Totales en Fila 27
    tot_row = 27
    ws3.cell(row=tot_row, column=1, value="TOTALES").font = bold_font
    ws3.cell(row=tot_row, column=1).alignment = align_center
    ws3.cell(row=tot_row, column=1).fill = total_fill
    ws3.cell(row=tot_row, column=1).border = border_total
    
    for c in range(2, 26):
        cell = ws3.cell(row=tot_row, column=c)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border_total
        
        # Sumas acumuladas para columnas clave
        if c == 4: # Cantidad
            cell.value = "=SUM(D12:D26)"
            cell.number_format = format_qty
            cell.alignment = align_right
        elif c == 9: # FOB Total
            cell.value = "=SUM(I12:I26)"
            cell.number_format = format_usd
            cell.alignment = align_right
        elif c == 10: # Peso Total
            cell.value = "=SUM(J12:J26)"
            cell.number_format = format_weight
            cell.alignment = align_right
        elif c == 11: # Vol Total
            cell.value = "=SUM(K12:K26)"
            cell.number_format = format_volume
            cell.alignment = align_right
        elif c == 12: # Flete
            cell.value = "=SUM(L12:L26)"
            cell.number_format = format_usd
            cell.alignment = align_right
        elif c == 13: # Seguro
            cell.value = "=SUM(M12:M26)"
            cell.number_format = format_usd
            cell.alignment = align_right
        elif c == 14: # CIF Total
            cell.value = "=SUM(N12:N26)"
            cell.number_format = format_usd
            cell.alignment = align_right
        elif c == 16: # Ad-Valorem
            cell.value = "=SUM(P12:P26)"
            cell.number_format = format_usd
            cell.alignment = align_right
        elif c == 17: # Gastos Locales
            cell.value = "=SUM(Q12:Q26)"
            cell.number_format = format_usd
            cell.alignment = align_right
        elif c == 18: # Costo Almacen
            cell.value = "=SUM(R12:R26)"
            cell.number_format = format_usd
            cell.alignment = align_right
        elif c == 23: # Utilidad
            cell.value = "=SUM(W12:W26)"
            cell.number_format = format_usd
            cell.alignment = align_right
        elif c == 24: # IGV+IPM
            cell.value = "=SUM(X12:X26)"
            cell.number_format = format_usd
            cell.alignment = align_right
        elif c == 25: # Percepción
            cell.value = "=SUM(Y12:Y26)"
            cell.number_format = format_usd
            cell.alignment = align_right
            
    ws3.row_dimensions[tot_row].height = 22

    # Ajustar anchos de columna de forma dinámica con un mínimo
    col_widths = {
        'A': 6, 'B': 13, 'C': 26, 'D': 13, 'E': 14, 'F': 13, 'G': 13, 'H': 10,
        'I': 15, 'J': 14, 'K': 14, 'L': 14, 'M': 14, 'N': 15, 'O': 14, 'P': 15,
        'Q': 14, 'R': 17, 'S': 16, 'T': 10, 'U': 20, 'V': 20, 'W': 16, 'X': 16, 'Y': 16
    }
    for col_letter, width in col_widths.items():
        ws3.column_dimensions[col_letter].width = width

    # Guardar archivo
    file_path = "Calculadora_Importacion.xlsx"
    wb.save(file_path)
    print(f"Libro de Excel creado exitosamente en: '{file_path}'")

if __name__ == "__main__":
    create_import_calculator()
